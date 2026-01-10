import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

# ==============================
# Configuration
# ==============================
INPUT_FILE = "sales.csv"
OUTPUT_FILE = "sales_report.xlsx"
SHEET_NAME = "Sales Report"

# ==============================
# Step 1: Read CSV Data
# ==============================
data = []

try:
    with open(INPUT_FILE, mode="r", newline="", encoding="utf-8") as file:
        reader = csv.DictReader(file)
        for row in reader:
            data.append(row)
except FileNotFoundError:
    print(f"❌ Error: '{INPUT_FILE}' not found. Please create it first.")
    exit()

if not data:
    print("❌ Error: CSV file is empty.")
    exit()

# ==============================
# Step 2: Create Workbook
# ==============================
wb = Workbook()
sheet = wb.active
sheet.title = SHEET_NAME

# ==============================
# Step 3: Write Headers
# ==============================
headers = ["Product", "Revenue", "Units"]
header_fill = PatternFill(start_color="0066CC", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")

for col_index, header in enumerate(headers, start=1):
    cell = sheet.cell(row=1, column=col_index, value=header)
    cell.font = header_font
    cell.fill = header_fill
    # Adjust column width for readability
    sheet.column_dimensions[get_column_letter(col_index)].width = 18

# ==============================
# Step 4: Write Data Rows
# ==============================
for row_index, row in enumerate(data, start=2):
    sheet.cell(row=row_index, column=1, value=row.get("Product"))
    sheet.cell(row=row_index, column=2, value=float(row.get("Revenue", 0)))
    sheet.cell(row=row_index, column=3, value=int(row.get("Units", 0)))

last_data_row = len(data) + 1

# ==============================
# Step 5: Add Totals
# ==============================
total_revenue = sum(float(row.get("Revenue", 0)) for row in data)
total_units = sum(int(row.get("Units", 0)) for row in data)

total_row = last_data_row + 2

sheet.cell(row=total_row, column=1, value="Total").font = Font(bold=True)
sheet.cell(row=total_row, column=2, value=total_revenue).font = Font(bold=True)
sheet.cell(row=total_row, column=3, value=total_units).font = Font(bold=True)

# ==============================
# Step 6: Create Chart
# ==============================
chart = BarChart()
chart.title = "Revenue by Product"
chart.y_axis.title = "Revenue ($)"
chart.x_axis.title = "Product"

# Data for the chart (Revenue column)
data_ref = Reference(
    sheet,
    min_col=2,
    min_row=1,
    max_row=last_data_row
)

# Categories for the chart (Product names)
category_ref = Reference(
    sheet,
    min_col=1,
    min_row=2,
    max_row=last_data_row
)

chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(category_ref)

# Place the chart starting at cell E5
sheet.add_chart(chart, "E5")

# ==============================
# Step 7: Save File
# ==============================
wb.save(OUTPUT_FILE)
print(f"✅ Excel sales report generated successfully: {OUTPUT_FILE}")